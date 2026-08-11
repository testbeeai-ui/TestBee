"""
Generate investor-facing TestBee (bytsiknhtcnlxwzgqkrd) public schema catalog:
  - Excel workbook
  - JSON for Cursor canvas
"""
from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# Exact row counts from pg_stat_user_tables (TestBee public)
ROWS = {
    "accepted_answer_payouts": 1,
    "admin_analytics_cache": 7,
    "admin_audit_log": 0,
    "admin_user_actions": 0,
    "ai_token_logs": 12310,
    "approved_emails": 8,
    "buddy_invites": 7,
    "cbse_mcq_chapters": 89,
    "cbse_mcq_community_share_rdm_claims": 0,
    "cbse_mcq_score_bonus_claims": 9,
    "class_exploration_sessions": 4,
    "classroom_assignment_completion_rdm_grants": 0,
    "classroom_assignment_responses": 0,
    "classroom_assignment_task_progress": 3,
    "classroom_generated_test_attempts": 0,
    "classroom_invite_batches": 1,
    "classroom_invite_recipients": 1,
    "classroom_join_requests": 8,
    "classroom_members": 30,
    "classroom_reviews": 3,
    "classroom_sections": 4,
    "classroom_subtopic_unlock_grants": 0,
    "classrooms": 8,
    "contact_messages": 5,
    "coupons": 3,
    "curriculum_chapters": 88,
    "curriculum_subtopics": 1199,
    "curriculum_topics": 386,
    "curriculum_units": 62,
    "daily_gauntlet_attempts": 11,
    "daily_reward_claims": 43,
    "dive_hub_progress": 12,
    "doubt_answer_reports": 2,
    "doubt_answers": 98,
    "doubt_saves": 6,
    "doubt_votes": 21,
    "doubts": 86,
    "edubite_brain_gym_progress": 6,
    "edubite_content_questions": 2994,
    "edubite_game_state": 11,
    "edubite_inspiration_blocks": 2,
    "edubite_inspiration_phenomena": 180,
    "edubite_inspiration_quotes": 185,
    "edubite_inspiration_role_models": 90,
    "edubite_monthly_challenge_enrollments": 3,
    "edubite_monthly_challenge_entries": 0,
    "edubite_pledge_reel_days": 360,
    "edubite_pledge_reel_slides": 1440,
    "edubite_puzzle_progress": 4,
    "edubite_rdm_rewards": 18,
    "edubite_reward_claims": 47,
    "edudeca_daily_attempts": 10,
    "edudeca_questions": 39,
    "edudeca_user_progress": 4,
    "episodic_memory": 0,
    "explorer_live_joins": 0,
    "gyan_bot_config": 1,
    "gyan_curriculum_nodes": 20,
    "inactive_day_penalties": 269,
    "learning_outcomes_questions": 390,
    "lessons_raw_post_boosts": 0,
    "lessons_raw_post_comments": 4,
    "lessons_raw_post_votes": 6,
    "lessons_raw_posts": 39,
    "live_class_ratings": 0,
    "live_class_slots": 5,
    "live_session_joins": 0,
    "live_sessions": 3,
    "magic_wall_basket_items": 20,
    "magic_wall_topic_attempts": 14,
    "mobile_push_tokens": 0,
    "mock_community_share_rdm_claims": 1,
    "mock_papers": 114,
    "mock_questions": 10520,
    "mock_rdm_bonus_attempts": 0,
    "mock_rdm_bonus_claims": 0,
    "mock_test_attempts": 25,
    "news_blog_posts": 9,
    "numerals_community_share_rdm_claims": 7,
    "numerals_formula_complete_rdm_claims": 1,
    "numerals_pack_complete_rdm_claims": 0,
    "past_paper_questions": 19853,
    "past_papers": 186,
    "platform_feedback_submissions": 0,
    "play_history": 399,
    "play_questions": 900,
    "posts": 64,
    "prep_calendar_day_activity": 72,
    "profile_academics": 7,
    "profile_achievements": 5,
    "profiles": 46,
    "quiz_community_share_rdm_claims": 6,
    "quiz_overall_complete_rdm_claims": 0,
    "quiz_set_complete_rdm_claims": 1,
    "rdm_config": 149,
    "refer_challenge_claims": 7,
    "referral_attributions": 1,
    "referral_weekly_bonuses": 0,
    "saved_questions": 38,
    "student_bits_attempts": 17,
    "student_events": 6943,
    "student_gyan_presence": 0,
    "student_learning_dwell": 1976,
    "student_learning_presence": 1,
    "student_lesson_mark_completions": 2,
    "student_section_history": 40,
    "student_site_presence": 6,
    "student_subtopic_engagement": 58,
    "study_buddies": 4,
    "study_streak_milestone_claims": 0,
    "subject_topic_chat_messages": 62,
    "subscription_coupons": 4,
    "subtopic_content": 1853,
    "teacher_generated_test_history": 11,
    "teacher_google_calendar_tokens": 2,
    "teacher_live_class_rdm_grants": 3,
    "teacher_motivation_rdm_grants": 0,
    "teacher_profile_details": 2,
    "teacher_section_schedule_rdm_grants": 14,
    "teacher_subscription_coupons": 0,
    "topic_content": 626,
    "topic_content_runs": 1033,
    "topic_quiz_advanced_rdm_attempts": 2,
    "transactional_email_logs": 102,
    "user_memory_profile": 0,
    "user_play_stats": 44,
    "user_roles": 5,
    "user_saved_items": 59,
    "user_study_day_totals": 194,
    "waitlist_submissions": 16,
}

# table -> (domain, purpose)
META: dict[str, tuple[str, str]] = {
    # Identity & Access
    "profiles": ("Identity & Access", "Core user profile for students and teachers (name, class, board, RDM wallet link fields)."),
    "user_roles": ("Identity & Access", "Role flags (student / teacher / admin) used for authorization."),
    "approved_emails": ("Identity & Access", "Allowlisted emails for gated access or early access cohorts."),
    "teacher_profile_details": ("Identity & Access", "Extended teacher identity, professional details, and optional KYC links."),
    "teacher_google_calendar_tokens": ("Identity & Access", "Secure Google Calendar OAuth tokens for teacher scheduling (server-only)."),
    "mobile_push_tokens": ("Identity & Access", "Device push-notification tokens for the mobile app."),
    # Curriculum & Content
    "curriculum_units": ("Curriculum & Content", "Top-level CBSE curriculum units (subject + class)."),
    "curriculum_chapters": ("Curriculum & Content", "Chapter catalog under each curriculum unit."),
    "curriculum_topics": ("Curriculum & Content", "Topics within each chapter."),
    "curriculum_subtopics": ("Curriculum & Content", "Finest learning nodes (subtopics) mapped for Dive and teaching."),
    "subtopic_content": ("Curriculum & Content", "Admin-managed theory content packs per subtopic (Dive lessons)."),
    "topic_content": ("Curriculum & Content", "Topic hub overview content, often AI-assisted and admin-editable."),
    "topic_content_runs": ("Curriculum & Content", "Audit history of topic content generation / regeneration runs."),
    "learning_outcomes_questions": ("Curriculum & Content", "Learning Outcomes MCQ packs per subtopic for Dive assessment."),
    "gyan_curriculum_nodes": ("Curriculum & Content", "CBSE cells that rotate automated Gyan++ bot questions."),
    "news_blog_posts": ("Curriculum & Content", "Public news / blog articles shown on marketing or in-app surfaces."),
    # Practice & Assessment
    "play_questions": ("Practice & Assessment", "Edublast Play question bank (academic + FunBrain style)."),
    "play_history": ("Practice & Assessment", "Per-answer Play history for adaptive rating and repeat control."),
    "user_play_stats": ("Practice & Assessment", "Per-user Play rating / streak stats by category."),
    "mock_papers": ("Practice & Assessment", "Catalog of timed mock exams (PYQ / NCERT style papers)."),
    "mock_questions": ("Practice & Assessment", "Question rows belonging to each mock paper."),
    "mock_test_attempts": ("Practice & Assessment", "Completed student mock sessions with scores and subject breakdown."),
    "past_papers": ("Practice & Assessment", "Past-paper catalog (previous-year papers)."),
    "past_paper_questions": ("Practice & Assessment", "Question rows for each past paper."),
    "cbse_mcq_chapters": ("Practice & Assessment", "CBSE chapter index for chapter-wise MCQ practice browser."),
    "student_bits_attempts": ("Practice & Assessment", "Student attempts for Dive Quiz (bits) practice sets."),
    "saved_questions": ("Practice & Assessment", "Bookmarked practice questions across mock / past-paper / static sources."),
    "user_saved_items": ("Practice & Assessment", "Generic saved learning items for later review."),
    "magic_wall_basket_items": ("Practice & Assessment", "Magic Wall reading-basket selections persisted per learner."),
    "magic_wall_topic_attempts": ("Practice & Assessment", "Learner attempts / progress against Magic Wall topics."),
    "daily_gauntlet_attempts": ("Practice & Assessment", "One Daily Gauntlet attempt per user per day for leaderboard."),
    "prep_calendar_day_activity": ("Practice & Assessment", "Prep calendar day-level activity completion markers."),
    "dive_hub_progress": ("Practice & Assessment", "Dive hub completion and Quiz / Numerals / Outcomes scores per subtopic."),
    # Community & Doubts
    "doubts": ("Community & Doubts", "Gyan++ student questions with voting and reward hooks."),
    "doubt_answers": ("Community & Doubts", "Answers to doubts; authors can accept a best answer."),
    "doubt_votes": ("Community & Doubts", "Up/down votes on doubts and answers (one vote per user per target)."),
    "doubt_saves": ("Community & Doubts", "Bookmarked doubts for later reading."),
    "doubt_answer_reports": ("Community & Doubts", "Abuse reports on answers; thresholds trigger penalties."),
    "accepted_answer_payouts": ("Community & Doubts", "Ledger of RDM paid when an answer is accepted (anti-farming / leaderboard)."),
    "gyan_bot_config": ("Community & Doubts", "Singleton config for automated Gyan++ student-persona posts."),
    "lessons_raw_posts": ("Community & Doubts", "Social feed posts on Lessons hub (non-Gyan++)."),
    "lessons_raw_post_comments": ("Community & Doubts", "Threaded comments on Lessons raw posts."),
    "lessons_raw_post_votes": ("Community & Doubts", "Votes on Lessons raw posts (+1 / -1)."),
    "lessons_raw_post_boosts": ("Community & Doubts", "Boost actions on Lessons raw posts (one per user per post)."),
    "posts": ("Community & Doubts", "Classroom / community posts including assignments and announcements."),
    "subject_topic_chat_messages": ("Community & Doubts", "Private topic chat transcripts for a learner (append-only)."),
    # Learning Buddies
    "study_buddies": ("Learning Buddies", "Active Learning Buddy pairs (symmetric rows; one active buddy per user)."),
    "buddy_invites": ("Learning Buddies", "WhatsApp-shareable Learning Buddy invite tokens."),
    "student_site_presence": ("Learning Buddies", "Latest site-wide focus heartbeat for buddy “right now” UI."),
    "student_learning_presence": ("Learning Buddies", "Latest subtopic panel a student is viewing (buddy presence)."),
    "student_gyan_presence": ("Learning Buddies", "Latest Gyan++ focus heartbeat for buddy presence."),
    # Classrooms & Teachers
    "classrooms": ("Classrooms & Teachers", "Teacher-owned classroom entities."),
    "classroom_sections": ("Classrooms & Teachers", "Sections / batches inside a classroom."),
    "classroom_members": ("Classrooms & Teachers", "Student memberships in classrooms."),
    "classroom_join_requests": ("Classrooms & Teachers", "Pending student join requests awaiting teacher approval."),
    "classroom_invite_batches": ("Classrooms & Teachers", "Bulk invite campaigns created by teachers."),
    "classroom_invite_recipients": ("Classrooms & Teachers", "Per-recipient status for classroom invite batches."),
    "classroom_reviews": ("Classrooms & Teachers", "Student reviews / ratings of classrooms."),
    "classroom_assignment_task_progress": ("Classrooms & Teachers", "Checklist completion for assignment tasks."),
    "classroom_assignment_responses": ("Classrooms & Teachers", "Optional student text/link submissions for assignment tasks."),
    "classroom_generated_test_attempts": ("Classrooms & Teachers", "Student attempts on teacher-generated MCQ tests."),
    "teacher_generated_test_history": ("Classrooms & Teachers", "History of tests generated by teachers for classrooms."),
    "classroom_subtopic_unlock_grants": ("Classrooms & Teachers", "Paid Concept Focus unlocks of a subtopic for targeted students."),
    "classroom_assignment_completion_rdm_grants": ("Classrooms & Teachers", "Escrowed teacher RDM paid when a student finishes an assignment on time."),
    "teacher_motivation_rdm_grants": ("Classrooms & Teachers", "Teacher-funded motivation bonuses paid on assignment completion."),
    "student_section_history": ("Classrooms & Teachers", "History of which classroom sections a student belonged to."),
    "class_exploration_sessions": ("Classrooms & Teachers", "Non-member timed classroom exploration windows."),
    "explorer_live_joins": ("Classrooms & Teachers", "Non-member live joins during exploration (time-capped)."),
    # Live Classes
    "live_sessions": ("Live Classes", "Conducted live class session records."),
    "live_session_joins": ("Live Classes", "Student joins to live sessions (first join may consume credits)."),
    "live_class_slots": ("Live Classes", "Booked live class occurrences used for delivery / quality rewards."),
    "live_class_ratings": ("Live Classes", "Per-student star ratings for a scheduled live occurrence."),
    "teacher_live_class_rdm_grants": ("Live Classes", "Idempotent RDM ledger when a live class is delivered."),
    "teacher_section_schedule_rdm_grants": ("Live Classes", "RDM grants for scheduled section occurrences (Path A delivery)."),
    # Economy & Rewards
    "rdm_config": ("Economy & Rewards", "Admin-editable RDM reward amounts and economy parameters."),
    "daily_reward_claims": ("Economy & Rewards", "Idempotent first-of-day reward claims (IST calendar day)."),
    "inactive_day_penalties": ("Economy & Rewards", "Ledger of inactivity penalties applied to wallets / streaks."),
    "study_streak_milestone_claims": ("Economy & Rewards", "Claims for study-streak milestone bonuses."),
    "mock_rdm_bonus_claims": ("Economy & Rewards", "Successful mock-test RDM bonus grants (capped per day / paper)."),
    "mock_rdm_bonus_attempts": ("Economy & Rewards", "Audit log of mock RDM bonus API successes and denials."),
    "mock_community_share_rdm_claims": ("Economy & Rewards", "RDM for verified community shares of mock results."),
    "quiz_set_complete_rdm_claims": ("Economy & Rewards", "RDM claims when a Quiz set is completed."),
    "quiz_overall_complete_rdm_claims": ("Economy & Rewards", "RDM claims when overall Quiz progress milestones complete."),
    "quiz_community_share_rdm_claims": ("Economy & Rewards", "RDM for sharing Quiz achievements to the community."),
    "numerals_formula_complete_rdm_claims": ("Economy & Rewards", "RDM when a Numerals formula practice item is completed."),
    "numerals_pack_complete_rdm_claims": ("Economy & Rewards", "RDM when a Numerals pack is fully completed."),
    "numerals_community_share_rdm_claims": ("Economy & Rewards", "RDM for sharing Numerals achievements."),
    "cbse_mcq_score_bonus_claims": ("Economy & Rewards", "Accuracy-tier RDM for finishing a CBSE chapter MCQ attempt."),
    "cbse_mcq_community_share_rdm_claims": ("Economy & Rewards", "RDM for community-sharing a CBSE chapter MCQ result."),
    "topic_quiz_advanced_rdm_attempts": ("Economy & Rewards", "Audit log for advanced topic-quiz daily RDM claims."),
    "coupons": ("Economy & Rewards", "General promotional / access coupon codes."),
    "subscription_coupons": ("Economy & Rewards", "Student subscription coupon codes and redemption metadata."),
    "teacher_subscription_coupons": ("Economy & Rewards", "Admin-issued teacher plan coupons (starter / pro tiers)."),
    # Growth & Referrals
    "referral_attributions": ("Growth & Referrals", "Canonical referee→referrer attribution (service-role writes only)."),
    "referral_weekly_bonuses": ("Growth & Referrals", "Guards weekly referrer RDM bonus (once per IST week)."),
    "refer_challenge_claims": ("Growth & Referrals", "Daily refer-challenge win/share reward claims."),
    "waitlist_submissions": ("Growth & Referrals", "Public waitlist form leads (student / teacher / parent)."),
    "contact_messages": ("Growth & Referrals", "Public Contact Us form submissions."),
    "platform_feedback_submissions": ("Growth & Referrals", "Structured in-app product feedback survey responses."),
    # Edubite Companion
    "edubite_content_questions": ("Edubite Companion", "Edubite DailyDose / FunBrain scheduled MCQ banks (isolated from Play)."),
    "edubite_game_state": ("Edubite Companion", "Per-user Edubite progress: streaks, habits, pledges, scores."),
    "edubite_brain_gym_progress": ("Edubite Companion", "Edubite Brain Gym mastery / progress JSON."),
    "edubite_puzzle_progress": ("Edubite Companion", "Edubite daily puzzle attempt progress."),
    "edubite_reward_claims": ("Edubite Companion", "Idempotent Edubite reward claim ledger."),
    "edubite_rdm_rewards": ("Edubite Companion", "Configurable Edubite RDM amounts (admin editable)."),
    "edubite_pledge_reel_days": ("Edubite Companion", "AI integrity pledge reel day themes (AM/PM packs)."),
    "edubite_pledge_reel_slides": ("Edubite Companion", "Individual slides for Edubite pledge reels."),
    "edubite_inspiration_quotes": ("Edubite Companion", "Rotating Inspiration quotes."),
    "edubite_inspiration_role_models": ("Edubite Companion", "Inspiration role-model cards."),
    "edubite_inspiration_phenomena": ("Edubite Companion", "Inspiration science / wonder phenomena cards."),
    "edubite_inspiration_blocks": ("Edubite Companion", "Inspiration content block definitions."),
    "edubite_monthly_challenge_enrollments": ("Edubite Companion", "Monthly Challenge enrollments per user."),
    "edubite_monthly_challenge_entries": ("Edubite Companion", "Day-level Monthly Challenge progress entries."),
    # EduDeca
    "edudeca_questions": ("EduDeca", "EduDeca question bank."),
    "edudeca_daily_attempts": ("EduDeca", "Daily EduDeca attempt records."),
    "edudeca_user_progress": ("EduDeca", "Longitudinal EduDeca progress per user."),
    # Analytics & Engagement
    "student_events": ("Analytics & Engagement", "Product analytics event stream (feature usage, funnels)."),
    "student_subtopic_engagement": ("Analytics & Engagement", "Per-subtopic engagement aggregates for a learner."),
    "student_lesson_mark_completions": ("Analytics & Engagement", "Lesson checklist ticks used for chapter/topic rollups."),
    "user_study_day_totals": ("Analytics & Engagement", "Daily study-time totals per user (IST day)."),
    "profile_academics": ("Analytics & Engagement", "Public-profile academic exam records (Class 10/12 etc.)."),
    "profile_achievements": ("Analytics & Engagement", "Public-profile achievements and competitions."),
    "student_learning_dwell": (
        "Analytics & Engagement",
        "Single table of active study-time samples (user, content, panel, tags, meta) with filter indexes.",
    ),
    # AI & Memory
    "ai_token_logs": ("AI & Memory", "Per-request AI usage telemetry (model, tokens, cost, backend)."),
    "episodic_memory": ("AI & Memory", "Vectorized episodic memory chunks for personalized recall."),
    "user_memory_profile": ("AI & Memory", "Canonical per-user memory profile JSON updated asynchronously."),
    # Admin & Ops
    "admin_analytics_cache": ("Admin & Ops", "Cached admin dashboard metrics for fast investor/ops views."),
    "admin_audit_log": ("Admin & Ops", "Immutable admin action audit trail."),
    "admin_user_actions": ("Admin & Ops", "Recorded admin interventions on user accounts."),
    "transactional_email_logs": ("Admin & Ops", "Outbound transactional email send log (welcome / login, IST day totals)."),
}

DOMAIN_ORDER = [
    "Identity & Access",
    "Curriculum & Content",
    "Practice & Assessment",
    "Community & Doubts",
    "Learning Buddies",
    "Classrooms & Teachers",
    "Live Classes",
    "Economy & Rewards",
    "Growth & Referrals",
    "Edubite Companion",
    "EduDeca",
    "Analytics & Engagement",
    "AI & Memory",
    "Admin & Ops",
]


def build_rows() -> list[dict]:
    missing = sorted(set(ROWS) - set(META))
    extra = sorted(set(META) - set(ROWS))
    if missing:
        raise SystemExit(f"Missing META for: {missing}")
    if extra:
        raise SystemExit(f"META without rows: {extra}")

    rows = []
    for i, name in enumerate(sorted(ROWS), 1):
        domain, purpose = META[name]
        rows.append(
            {
                "no": i,
                "table": name,
                "domain": domain,
                "purpose": purpose,
                "rows": ROWS[name],
                "schema": "public",
                "project": "TestBee",
                "project_id": "bytsiknhtcnlxwzgqkrd",
            }
        )
    # renumber after domain sort for nicer investor sheet
    rows.sort(key=lambda r: (DOMAIN_ORDER.index(r["domain"]), r["table"]))
    for i, r in enumerate(rows, 1):
        r["no"] = i
    return rows


def style_header(cell, fill: PatternFill):
    cell.font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
    cell.fill = fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def write_excel(rows: list[dict], path: Path):
    wb = Workbook()

    # --- Cover ---
    cover = wb.active
    cover.title = "Cover"
    navy = PatternFill("solid", fgColor="0F2744")
    teal = PatternFill("solid", fgColor="0F766E")
    soft = PatternFill("solid", fgColor="F0F7F6")
    thin = Border(
        left=Side(style="thin", color="D0D7DE"),
        right=Side(style="thin", color="D0D7DE"),
        top=Side(style="thin", color="D0D7DE"),
        bottom=Side(style="thin", color="D0D7DE"),
    )

    cover["A1"] = "EduBlast / TestBee"
    cover["A1"].font = Font(name="Calibri", bold=True, size=22, color="0F2744")
    cover["A2"] = "Supabase Data Model — Investor Catalog"
    cover["A2"].font = Font(name="Calibri", size=14, color="0F766E")
    cover["A4"] = "Project"
    cover["B4"] = "TestBee"
    cover["A5"] = "Project ID"
    cover["B5"] = "bytsiknhtcnlxwzgqkrd"
    cover["A6"] = "Schema"
    cover["B6"] = "public"
    cover["A7"] = "Tables cataloged"
    cover["B7"] = len(rows)
    cover["A8"] = "Total live rows (approx)"
    cover["B8"] = sum(r["rows"] for r in rows)
    cover["A9"] = "Source"
    cover["B9"] = "Live Supabase metadata + table comments (Aug 2026)"
    cover["A11"] = (
        "This workbook lists every public table in the production TestBee database "
        "with a short business-purpose description suitable for diligence. "
        "TestBee RAG is excluded by design."
    )
    cover["A11"].alignment = Alignment(wrap_text=True, vertical="top")
    cover.merge_cells("A11:F13")
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 55
    for r in range(4, 10):
        cover[f"A{r}"].font = Font(name="Calibri", bold=True, color="0F2744")
        cover[f"A{r}"].fill = soft
        cover[f"B{r}"].fill = soft

    # --- Domain summary ---
    summary = wb.create_sheet("Domain Summary")
    headers = ["Domain", "Tables", "Live Rows", "What investors should know"]
    domain_blurbs = {
        "Identity & Access": "Users, roles, and secure credentials that gate the platform.",
        "Curriculum & Content": "CBSE tree + Dive theory / Learning Outcomes content IP.",
        "Practice & Assessment": "Question banks and attempt telemetry (Play, Mock, PYQ).",
        "Community & Doubts": "Gyan++ Q&A and social lesson surfaces that drive engagement.",
        "Learning Buddies": "Peer accountability graph and live presence signals.",
        "Classrooms & Teachers": "Teacher portal: classes, assignments, unlocks, invites.",
        "Live Classes": "Live delivery inventory, joins, ratings, and teacher payouts.",
        "Economy & Rewards": "RDM ledger rules and idempotent claim tables (anti-abuse).",
        "Growth & Referrals": "Acquisition, referrals, waitlist, and feedback loops.",
        "Edubite Companion": "Sibling daily-learning product data (isolated from Edublast Play).",
        "EduDeca": "EduDeca product question and progress tables.",
        "Analytics & Engagement": "Events, dwell partitions, and study-time aggregates.",
        "AI & Memory": "Token cost telemetry and personalization memory stores.",
        "Admin & Ops": "Admin caches, audits, and transactional email logs.",
    }
    for col, h in enumerate(headers, 1):
        cell = summary.cell(1, col, h)
        style_header(cell, navy)
    summary.row_dimensions[1].height = 28

    for i, domain in enumerate(DOMAIN_ORDER, 2):
        subset = [r for r in rows if r["domain"] == domain]
        summary.cell(i, 1, domain).font = Font(name="Calibri", bold=True)
        summary.cell(i, 2, len(subset))
        summary.cell(i, 3, sum(r["rows"] for r in subset))
        summary.cell(i, 4, domain_blurbs[domain])
        for c in range(1, 5):
            summary.cell(i, c).border = thin
            summary.cell(i, c).alignment = Alignment(vertical="center", wrap_text=True)
            if i % 2 == 0:
                summary.cell(i, c).fill = soft
    summary.column_dimensions["A"].width = 28
    summary.column_dimensions["B"].width = 12
    summary.column_dimensions["C"].width = 14
    summary.column_dimensions["D"].width = 70
    for i in range(2, 2 + len(DOMAIN_ORDER)):
        summary.row_dimensions[i].height = 36

    # --- All Tables ---
    all_sheet = wb.create_sheet("All Tables", 1)
    cols = ["#", "Domain", "Table", "Purpose (investor)", "Live Rows", "Schema", "RLS"]
    for col, h in enumerate(cols, 1):
        cell = all_sheet.cell(1, col, h)
        style_header(cell, teal)
    all_sheet.row_dimensions[1].height = 30
    all_sheet.freeze_panes = "A2"
    all_sheet.auto_filter.ref = f"A1:G{len(rows) + 1}"

    for i, r in enumerate(rows, 2):
        values = [r["no"], r["domain"], r["table"], r["purpose"], r["rows"], "public", "Enabled"]
        for c, v in enumerate(values, 1):
            cell = all_sheet.cell(i, c, v)
            cell.border = thin
            cell.alignment = Alignment(vertical="center", wrap_text=(c in (2, 4)))
            cell.font = Font(name="Calibri", size=10)
            if i % 2 == 0:
                cell.fill = soft
            if c == 3:
                cell.font = Font(name="Consolas", size=10, color="0F2744")
            if c == 5:
                cell.alignment = Alignment(horizontal="right", vertical="center")
        all_sheet.row_dimensions[i].height = 42

    widths = [6, 24, 42, 78, 12, 10, 10]
    for i, w in enumerate(widths, 1):
        all_sheet.column_dimensions[get_column_letter(i)].width = w

    # Per-domain sheets
    for domain in DOMAIN_ORDER:
        subset = [r for r in rows if r["domain"] == domain]
        title = domain[:28]
        ws = wb.create_sheet(title)
        for col, h in enumerate(["#", "Table", "Purpose", "Live Rows"], 1):
            style_header(ws.cell(1, col, h), navy)
        ws.freeze_panes = "A2"
        for i, r in enumerate(subset, 2):
            ws.cell(i, 1, i - 1)
            ws.cell(i, 2, r["table"]).font = Font(name="Consolas", size=10)
            ws.cell(i, 3, r["purpose"])
            ws.cell(i, 4, r["rows"])
            for c in range(1, 5):
                ws.cell(i, c).border = thin
                ws.cell(i, c).alignment = Alignment(vertical="center", wrap_text=(c == 3))
                if i % 2 == 0:
                    ws.cell(i, c).fill = soft
            ws.row_dimensions[i].height = 40
        ws.column_dimensions["A"].width = 6
        ws.column_dimensions["B"].width = 42
        ws.column_dimensions["C"].width = 80
        ws.column_dimensions["D"].width = 12

    wb.save(path)
    print("Excel:", path)


def main():
    rows = build_rows()
    out_dir = Path(r"C:\Users\rentk\Desktop\Edublast")
    xlsx = out_dir / "TestBee_Supabase_Table_Catalog_Investor_v2.xlsx"
    write_excel(rows, xlsx)

    json_path = out_dir / "Web" / "tmp-testbee-table-catalog.json"
    json_path.write_text(json.dumps(rows, indent=2), encoding="utf-8")
    print("JSON:", json_path)
    print("tables", len(rows), "total_rows", sum(r["rows"] for r in rows))
    for d in DOMAIN_ORDER:
        subset = [r for r in rows if r["domain"] == d]
        print(f"  {d}: {len(subset)} tables")


if __name__ == "__main__":
    main()
